"""
build_inventory.py — generate a workspace's master Excel inventory.
Outputs to output/<workspace-slug>/workflow_master_inventory.xlsx.
"""
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from parser import Workspace, list_workspaces, OUTPUT_DIR

# ── styles ─────────────────────────────────────────────────────────
HFONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HFILL  = PatternFill("solid", start_color="2C3E50")
PKFILL = PatternFill("solid", start_color="EAF3DE")
FKFILL = PatternFill("solid", start_color="E6F1FB")
BFONT  = Font(name="Arial", size=10)
NFONT  = Font(name="Arial", size=10, italic=True, color="5F5E5A")
TFONT  = Font(name="Arial", size=14, bold=True)
SFONT  = Font(name="Arial", size=11, bold=True)
THIN   = Side(border_style="thin", color="B4B2A9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP   = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR    = Alignment(horizontal="center", vertical="center")

def sheet(ws, title, cols, rows, pk=None, fks=None):
    ws.sheet_view.showGridLines = False
    fks = fks or []
    ws["A1"] = title; ws["A1"].font = TFONT; ws.row_dimensions[1].height = 22
    for i,(h,w,_n) in enumerate(cols,1):
        c = ws.cell(row=3,column=i,value=h)
        c.font=HFONT; c.fill=HFILL; c.alignment=CTR; c.border=BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    for i,(h,_w,n) in enumerate(cols,1):
        c = ws.cell(row=4,column=i,value=n)
        c.font=NFONT; c.alignment=WRAP; c.border=BORDER
        if h == pk: c.fill = PKFILL
        elif h in fks: c.fill = FKFILL
    ws.row_dimensions[4].height = 36
    for r_idx,row in enumerate(rows, start=5):
        for c_idx,(h,_w,_n) in enumerate(cols,1):
            cell = ws.cell(row=r_idx,column=c_idx,value=row.get(h,""))
            cell.font = BFONT; cell.alignment = WRAP; cell.border = BORDER
        ws.row_dimensions[r_idx].height = 22
    ws.freeze_panes = "A5"
    if rows:
        ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{4+len(rows)}"

# ── load manual context (business processes, callsigns, etc.) ─────
def load_business_processes(ws):
    bp = ws.business_processes()
    if bp:
        return bp
    return [
        {"ProcessID":"INTAKE",  "ProcessName":"Customer Intake & Account Setup",
         "OwnerArea":"Energy Programs", "Description":"Initial customer info capture, eligibility.","Notes":""},
        {"ProcessID":"ENROLL",  "ProcessName":"Enrollment", "OwnerArea":"Energy Programs",
         "Description":"Customer signs enrollment; 310 record created.","Notes":""},
        {"ProcessID":"ASSESS",  "ProcessName":"Pre-Install Assessment","OwnerArea":"Energy Programs",
         "Description":"Home assessment; 315 record captures findings.","Notes":""},
        {"ProcessID":"INSTALL", "ProcessName":"Installation","OwnerArea":"Energy Programs",
         "Description":"Measures installed; 325 record captures execution.","Notes":""},
        {"ProcessID":"INSPECT", "ProcessName":"Inspection / QC","OwnerArea":"Compliance",
         "Description":"Post-install inspection. 395 record holds outcome.","Notes":""},
        {"ProcessID":"INVOICE", "ProcessName":"Invoicing & Payment","OwnerArea":"Operations",
         "Description":"Contractor invoicing; payment status; chargebacks.","Notes":""},
    ]

def build(workspace):
    data = workspace.discover()
    workspace_name = data["workspaceName"]
    wb = Workbook()

    # README
    ws = wb.active; ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 120
    ws["A1"] = f"Workflow & Form Master Inventory · {workspace_name}"
    ws["A1"].font = Font(name="Arial", size=16, bold=True)
    blocks = [
        ("", None),
        ("Auto-generated. Do not edit by hand — re-run scripts/build_inventory.py to refresh.", NFONT),
        ("", None),
        ("Sheet relationships", SFONT),
        ("FOUNDATION (parsed from the workspace export and/or data/forms/*.json):", SFONT),
        ("  Forms · Fields · FormRelationships · ReferencedDataFields", BFONT),
        ("WORKFLOWS (embedded in the workspace export and/or data/workflows/*.json):", SFONT),
        ("  Workflows · Triggers · Steps · Actions · WorkflowFieldUsage", BFONT),
        ("CONTEXT (manual, see data/manual/):", SFONT),
        ("  BusinessProcesses · QueryExamples", BFONT),
    ]
    r = 2
    for text, font in blocks:
        c = ws.cell(row=r, column=1, value=text)
        if font: c.font = font
        c.alignment = WRAP
        r += 1

    # Forms
    forms_rows = [{
        "FormName": f["name"], "Workspace": workspace_name, "Role": f["role"],
        "ParentForm": f.get("subformOf", ""),
        "FieldCount": f["fieldCount"],
        "Description": f.get("description", ""),
        "DuplicateRules": f.get("duplicateRules", ""),
        "SavedFilters": (lambda sf: f"{len(sf)}: " + ", ".join(sf) if sf else "")(f.get("savedFilters") or []),
        "SourceFile": f.get("sourceFile") or "(no JSON)",
        "Notes": "",
    } for f in data["forms"]]
    sheet(wb.create_sheet("Forms"), "Forms · master inventory",
          [("FormName",30,"PK · stable name"), ("Workspace",32,"Workspace"),
           ("Role",12,"Hub/Spoke/Lookup/Subform"),
           ("ParentForm",28,"Containing form (subforms only)"),
           ("FieldCount",12,"Total data fields"),
           ("Description",44,"From the workspace export"),
           ("DuplicateRules",44,"Duplicate-response match rules"),
           ("SavedFilters",40,"Count + names of saved views"),
           ("SourceFile",50,"Source JSON file (or empty if no profile yet)"),
           ("Notes",30,"Free-form")],
          forms_rows, pk="FormName")

    # Fields
    field_rows = []
    for form_name, flds in data["fields"].items():
        for fld in flds:
            dep = fld.get("dependsOn", {}) or {}
            field_rows.append({
                "FieldKey": f"{form_name}::{fld['name']}",
                "FormName": form_name, "FieldName": fld["name"], "Label": fld["label"],
                "DataType": fld["type"], "ComponentType": fld["component"],
                "Required": fld["required"], "Hidden": fld["hidden"], "Enabled": fld["enabled"],
                "IsRelationship": "Yes" if fld["component"]=="FormRelationshipInput" else "",
                "IsRefData": "Yes" if fld["component"]=="FormRelationshipReferenceDataInput" else "",
                "RelatedForm": fld["relatedForm"], "RelatedField": fld["relatedField"],
                "ViaRelationship": fld["via"],
                "Validator": fld.get("validator",""),
                "Formula": (fld.get("formula","") or "")[:500],
                "FilterCondition": ", ".join(dep.get("filter",[])) or fld.get("filter",""),
                "VisibilityCondition": ", ".join(dep.get("visibility",[])) or fld.get("visibility",""),
                "DefaultValue": fld.get("defaultValue",""),
                "AutoIncrement": fld.get("autoIncrement",""),
                "DependsOn": ", ".join(fld.get("dependsOnAll",[])),
            })
    sheet(wb.create_sheet("Fields"), "Fields · every field on every form",
          [("FieldKey",38,"PK · FormName::FieldName"), ("FormName",30,"FK → Forms"),
           ("FieldName",28,"System name"), ("Label",32,"Display label"),
           ("DataType",14,"Text/Decimal/Date/etc."), ("ComponentType",30,"Input control type"),
           ("Required",10,"Yes/No"), ("Hidden",10,"Yes/No"), ("Enabled",10,"Yes/No"),
           ("IsRelationship",14,"Yes if defines link"), ("IsRefData",14,"Yes if pulls through link"),
           ("RelatedForm",30,"Target form"), ("RelatedField",24,"Pulled field"),
           ("ViaRelationship",24,"Which rel field"),
           ("Validator",16,"Field-type validator"), ("Formula",48,"Computed-field formula"),
           ("FilterCondition",28,"Picklist filter depends on"),
           ("VisibilityCondition",28,"Visible-when depends on"),
           ("DefaultValue",18,"Static default"),
           ("AutoIncrement",24,"Prefix + counter start, if auto-numbered"),
           ("DependsOn",36,"All same-form fields this field references")],
          field_rows, pk="FieldKey", fks=["FormName"])

    # FormRelationships
    sheet(wb.create_sheet("FormRelationships"),
          "FormRelationships · declared links",
          [("SourceForm",30,"FK → Forms"), ("RelationField",28,"Field holding the link"),
           ("RelationLabel",28,"Display label"), ("TargetForm",30,"FK → Forms"),
           ("TargetMatchField",28,"Match field on target")],
          [{"SourceForm":r["source"], "RelationField":r["via"], "RelationLabel":r["label"],
            "TargetForm":r["target"], "TargetMatchField":r.get("targetMatchField","")}
           for r in data["relationships"]],
          fks=["SourceForm","TargetForm"])

    # ReferencedDataFields
    sheet(wb.create_sheet("ReferencedDataFields"),
          "ReferencedDataFields · cross-form pulls",
          [("DestinationForm",30,"FK → Forms"), ("DestinationField",28,"Local display field"),
           ("DestinationLabel",28,"Display label"), ("ViaRelationship",24,"Relationship used"),
           ("SourceField",24,"Field pulled"), ("DataType",12,"Type")],
          [{"DestinationForm":r["destForm"], "DestinationField":r["destField"],
            "DestinationLabel":r["destLabel"], "ViaRelationship":r["via"],
            "SourceField":r["sourceField"], "DataType":r["dataType"]}
           for r in data["refPulls"]],
          fks=["DestinationForm"])

    # Workflows
    wf_rows = [{
        "Callsign": w["callsign"], "WorkflowName": w["name"],
        "Description": w["description"],
        "BusinessProcess": "", "Workspace": workspace_name, "Owner": "",
        # Configured-but-not-running automations must be flagged, not rendered
        # like live ones — Status comes from the export's enabled flag.
        "Status": "Active" if w.get("enabled", True) else "Disabled",
        "Environment": "Staging", "Criticality": "",
        "Staging_GUID": w["staging_guid"], "Prod_GUID": "",
        "SourceFile": w["sourceFile"], "Notes": "",
    } for w in data["workflows"]]
    sheet(wb.create_sheet("Workflows"), "Workflows · automation master",
          [("Callsign",14,"PK · short alias"), ("WorkflowName",24,"Display name"),
           ("Description",40,"Plain-English summary"),
           ("BusinessProcess",20,"FK → BusinessProcesses"),
           ("Workspace",28,"Workspace"), ("Owner",24,"Name/email"),
           ("Status",12,"Active/Disabled"), ("Environment",14,"Staging/Prod"),
           ("Criticality",12,"High/Med/Low"),
           ("Staging_GUID",38,"Workflow ID in staging"), ("Prod_GUID",38,"ID in prod"),
           ("SourceFile",40,"Source JSON"), ("Notes",40,"Free-form")],
          wf_rows, pk="Callsign", fks=["BusinessProcess"])

    # Triggers
    trig_rows = []
    for w in data["workflows"]:
        if not w["trigger"]: continue
        t = w["trigger"]
        trig_rows.append({
            "TriggerName": "Start", "Callsign": w["callsign"],
            "TriggerType": t["type"], "SourceForm": t["form"],
            "SourceWorkspace": t["workspace"], "DatabaseAction": t["databaseAction"],
            "ActionTiming": t["timing"], "ConditionMode": t["conditionMode"],
            "ConditionSummary": t["condition"], "CronExpression": t["cron"] or "",
            "TimeZone": t["timezone"] or "", "MonitoredFields": "",
        })
    sheet(wb.create_sheet("Triggers"), "Triggers · what fires each workflow",
          [("TriggerName",16,"Display name"), ("Callsign",14,"FK → Workflows"),
           ("TriggerType",16,"FormResponse/Scheduled/Manual"), ("SourceForm",30,"FK → Forms"),
           ("SourceWorkspace",28,"Workspace"), ("DatabaseAction",14,"Create/Update/Delete"),
           ("ActionTiming",16,"Pre/Post Processing"), ("ConditionMode",14,"None/Basic/Expr"),
           ("ConditionSummary",44,"Plain English"), ("CronExpression",16,"If scheduled"),
           ("TimeZone",22,"If scheduled"), ("MonitoredFields",30,"If field-change mode")],
          trig_rows, fks=["Callsign","SourceForm"])

    # Actions
    act_rows = []
    for w in data["workflows"]:
        for a in w["actions"]:
            cat = "FormOps" if "FormResponse" in a["type"] else \
                  "Communication" if "Email" in a["type"] or "SMS" in a["type"] else \
                  "DocGen" if "PDF" in a["type"] else \
                  "Integration" if "WebHook" in a["type"] or "API" in a["type"] else "Other"
            act_rows.append({
                "Callsign": w["callsign"], "StepName": a["stepName"],
                "ActionDisplayName": a["name"], "ActionType": a["type"], "Category": cat,
                "TargetForm": a["targetForm"], "TargetWorkspace": a["targetWorkspace"],
                "DuplicateMatchPolicy": a["duplicatePolicy"],
                "ResolutionType": a["resolutionType"], "MatchFilterSummary": a["matchOn"],
                "ContinueOnError": "Yes" if a["continueOnError"] else "No",
                "Summary": "",
            })
    sheet(wb.create_sheet("Actions"), "Actions · what each step does",
          [("Callsign",14,"FK → Workflows"), ("StepName",24,"Step"),
           ("ActionDisplayName",24,"Human name"), ("ActionType",28,"BuiltIn.*"),
           ("Category",14,"Comm/FormOps/DocGen/Integration"),
           ("TargetForm",30,"FK → Forms"), ("TargetWorkspace",28,"Workspace"),
           ("DuplicateMatchPolicy",16,"Skip/Update/Create"),
           ("ResolutionType",16,"Filter/Expr/None"), ("MatchFilterSummary",44,"Plain English"),
           ("ContinueOnError",14,"Yes/No"), ("Summary",44,"Plain English")],
          act_rows, fks=["Callsign","TargetForm"])

    # WorkflowFieldUsage
    wfu_rows = []
    for w in data["workflows"]:
        for u in w["fieldUsage"]:
            wfu_rows.append({
                "Callsign": w["callsign"], "StepName": u.get("stepName",""),
                "FormName": u["form"], "FieldName": u["field"],
                "Direction": u["direction"], "UsageContext": u["context"],
                "Notes": "",
            })
    sheet(wb.create_sheet("WorkflowFieldUsage"),
          "WorkflowFieldUsage · every field a workflow touches",
          [("Callsign",14,"FK → Workflows"), ("StepName",24,"Step (or blank if trigger)"),
           ("FormName",30,"FK → Forms"), ("FieldName",24,"Field"),
           ("Direction",12,"Read/Write/Match/Condition"),
           ("UsageContext",38,"Where in the workflow"), ("Notes",30,"Free-form")],
          wfu_rows, fks=["Callsign","FormName"])

    # BusinessProcesses
    bp_rows = load_business_processes(workspace)
    sheet(wb.create_sheet("BusinessProcesses"),
          "BusinessProcesses · real-world function tagging",
          [("ProcessID",14,"PK · slug used in Workflows.BusinessProcess"),
           ("ProcessName",32,"Display name"), ("OwnerArea",22,"Functional owner"),
           ("Description",55,"What this covers"), ("Notes",40,"Free-form")],
          bp_rows, pk="ProcessID")

    # Reorder
    desired = ["README","Forms","Fields","FormRelationships","ReferencedDataFields",
               "Workflows","Triggers","Steps" if False else "Triggers",
               "Actions","WorkflowFieldUsage","BusinessProcesses"]
    desired = [n for n in ["README","Forms","Fields","FormRelationships","ReferencedDataFields",
                           "Workflows","Triggers","Actions","WorkflowFieldUsage","BusinessProcesses"]
               if n in wb.sheetnames]
    wb._sheets = [wb[n] for n in desired]

    out_dir = OUTPUT_DIR / workspace.slug
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "workflow_master_inventory.xlsx"
    wb.save(out)
    print(f"  Forms: {len(data['forms'])}  Fields: {sum(len(v) for v in data['fields'].values())}  "
          f"Relationships: {len(data['relationships'])}  Workflows: {len(data['workflows'])}")
    print(f"  Saved -> {out.relative_to(OUTPUT_DIR.parent)}")

if __name__ == "__main__":
    for slug in list_workspaces():
        print(f"[{slug}]")
        build(Workspace(slug))
