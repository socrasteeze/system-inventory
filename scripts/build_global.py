"""
build_global.py — cross-workspace aggregator.

Merges every workspace into a single view for impact analysis and
duplicate-flow detection. Outputs:
    output/global/cross-workspace-inventory.xlsx
    output/global/global-explorer.html
"""
import sys, json
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from parser import discover_all, OUTPUT_DIR
from build_inventory import sheet, TFONT, NFONT, SFONT, BFONT, WRAP

GLOBAL_DIR = OUTPUT_DIR / "global"
TEMPLATE = (Path(__file__).resolve().parent / "global_template.html").read_text(encoding="utf-8")


def _flow_signature(wf):
    """Normalized fingerprint of a workflow: trigger action + the set of target forms.

    Two workflows in different workspaces that fire on the same database action and
    touch the same target forms share a signature — the signal for a duplicated flow.
    """
    action = (wf["trigger"]["databaseAction"] if wf["trigger"] else "") or "(none)"
    targets = sorted({a["targetForm"] for a in wf["actions"] if a["targetForm"]})
    return f"{action} -> {', '.join(targets) if targets else '(no target)'}"


def aggregate():
    """Discover all workspaces and compute the cross-workspace rollups."""
    discovered = discover_all()

    workspaces = []
    all_forms = []        # rows with Workspace column
    all_workflows = []
    forms_by_name = {}     # display name -> set of slugs
    form_roles = {}        # display name -> set of roles
    flows_by_sig = {}      # signature -> list of {slug, callsign, name}

    for slug, data in discovered.items():
        name = data["workspaceName"]
        workspaces.append({
            "slug": slug, "name": name,
            "forms": len(data["forms"]),
            "fields": sum(len(v) for v in data["fields"].values()),
            "relationships": len(data["relationships"]),
            "workflows": len(data["workflows"]),
        })
        for f in data["forms"]:
            all_forms.append({"slug": slug, "name": name, **f})
            forms_by_name.setdefault(f["name"], set()).add(slug)
            form_roles.setdefault(f["name"], set()).add(f["role"])
        for w in data["workflows"]:
            targets = sorted({a["targetForm"] for a in w["actions"] if a["targetForm"]})
            all_workflows.append({
                "slug": slug, "name": name, "callsign": w["callsign"],
                "workflow": w["name"],
                "triggerForm": w["trigger"]["form"] if w["trigger"] else "",
                "triggerAction": w["trigger"]["databaseAction"] if w["trigger"] else "",
                "targets": targets,
                "writes": sum(1 for u in w["fieldUsage"] if u["direction"] == "Write"),
                "enabled": w.get("enabled", True),
            })
            sig = _flow_signature(w)
            flows_by_sig.setdefault(sig, []).append(
                {"slug": slug, "callsign": w["callsign"], "name": w["name"]})

    # A collision is one display name carried by two or more workspaces.
    collisions = sorted(
        ({"name": n, "slugs": sorted(s), "roles": sorted(form_roles.get(n, set()))}
         for n, s in forms_by_name.items() if len(s) >= 2),
        key=lambda c: (-len(c["slugs"]), c["name"]))

    # A duplicate flow is one signature shared by two or more workflows.
    duplicate_flows = sorted(
        ({"signature": sig, "members": members} for sig, members in flows_by_sig.items()
         if len(members) >= 2),
        key=lambda d: (-len(d["members"]), d["signature"]))

    return {
        "discovered": discovered,
        "workspaces": sorted(workspaces, key=lambda w: w["slug"]),
        "allForms": all_forms,
        "allWorkflows": all_workflows,
        "collisions": collisions,
        "duplicateFlows": duplicate_flows,
    }


def _build_excel(agg):
    wb = Workbook()

    ws = wb.active; ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 120
    ws["A1"] = "Cross-Workspace Inventory"
    ws["A1"].font = TFONT
    blocks = [
        ("", None),
        ("Auto-generated. Do not edit by hand — re-run scripts/build_global.py to refresh.", NFONT),
        ("", None),
        ("Sheets", SFONT),
        ("  Workspaces · one row per workspace", BFONT),
        ("  AllForms · every form across all workspaces", BFONT),
        ("  AllWorkflows · every workflow across all workspaces", BFONT),
        ("  FormNameCollisions · form names shared by 2+ workspaces (rename-impact targets)", BFONT),
        ("  DuplicateFlows · workflows with the same trigger/target signature (consolidation targets)", BFONT),
    ]
    r = 2
    for text, font in blocks:
        c = ws.cell(row=r, column=1, value=text)
        if font: c.font = font
        c.alignment = WRAP
        r += 1

    sheet(wb.create_sheet("Workspaces"), "Workspaces · roll-up",
          [("Slug",18,"Folder under data/"), ("DisplayName",36,"Workspace name"),
           ("Forms",10,"Form count"), ("Fields",10,"Total fields"),
           ("Relationships",14,"Declared links"), ("Workflows",12,"Workflow count")],
          [{"Slug":w["slug"], "DisplayName":w["name"], "Forms":w["forms"],
            "Fields":w["fields"], "Relationships":w["relationships"], "Workflows":w["workflows"]}
           for w in agg["workspaces"]],
          pk="Slug")

    sheet(wb.create_sheet("AllForms"), "AllForms · every form, every workspace",
          [("Workspace",18,"FK -> Workspaces"), ("FormName",30,"Display name"),
           ("Role",12,"Hub/Spoke/Lookup"), ("FieldCount",12,"Total fields"),
           ("SourceFile",50,"Source JSON (or empty if stubbed)")],
          [{"Workspace":f["slug"], "FormName":f["name"], "Role":f["role"],
            "FieldCount":f["fieldCount"], "SourceFile":f.get("sourceFile") or "(no JSON)"}
           for f in agg["allForms"]],
          fks=["Workspace"])

    sheet(wb.create_sheet("AllWorkflows"), "AllWorkflows · every workflow, every workspace",
          [("Workspace",18,"FK -> Workspaces"), ("Callsign",14,"Short alias"),
           ("WorkflowName",26,"Display name"), ("Status",12,"Active/Disabled"),
           ("TriggerForm",28,"Form that fires it"),
           ("TriggerAction",14,"Create/Update/Delete"), ("TargetForms",40,"Forms it writes"),
           ("Writes",10,"Field writes")],
          [{"Workspace":w["slug"], "Callsign":w["callsign"], "WorkflowName":w["workflow"],
            "Status":"Active" if w.get("enabled", True) else "Disabled",
            "TriggerForm":w["triggerForm"], "TriggerAction":w["triggerAction"],
            "TargetForms":", ".join(w["targets"]), "Writes":w["writes"]}
           for w in agg["allWorkflows"]],
          fks=["Workspace"])

    sheet(wb.create_sheet("FormNameCollisions"),
          "FormNameCollisions · same form name in 2+ workspaces",
          [("FormName",30,"Display name"), ("WorkspaceCount",16,"How many workspaces"),
           ("Workspaces",46,"Which workspaces"), ("Roles",24,"Role in each")],
          [{"FormName":c["name"], "WorkspaceCount":len(c["slugs"]),
            "Workspaces":", ".join(c["slugs"]), "Roles":", ".join(c["roles"])}
           for c in agg["collisions"]])

    dup_rows = []
    for d in agg["duplicateFlows"]:
        dup_rows.append({
            "Signature": d["signature"],
            "Count": len(d["members"]),
            "Workspaces": ", ".join(sorted({m["slug"] for m in d["members"]})),
            "Workflows": ", ".join(f"{m['slug']}/{m['callsign']}" for m in d["members"]),
        })
    sheet(wb.create_sheet("DuplicateFlows"),
          "DuplicateFlows · shared trigger/target signature",
          [("Signature",54,"Trigger action -> target forms"), ("Count",10,"Matching workflows"),
           ("Workspaces",30,"Workspaces involved"), ("Workflows",46,"slug/callsign")],
          dup_rows)

    GLOBAL_DIR.mkdir(parents=True, exist_ok=True)
    out = GLOBAL_DIR / "cross-workspace-inventory.xlsx"
    wb.save(out)
    print(f"  Saved -> {out.relative_to(OUTPUT_DIR.parent)}")


def _build_html(agg):
    def fid(slug, name): return f"{slug}::{name}"

    forms, workflows, relationships, wf_edges = [], [], [], []
    for slug, data in agg["discovered"].items():
        form_names = {f["name"] for f in data["forms"]}
        for f in data["forms"]:
            forms.append({"id": fid(slug, f["name"]), "name": f["name"], "slug": slug,
                          "role": f["role"], "fieldCount": f["fieldCount"]})
        # One edge per direction; a direction carrying many relationship fields
        # (e.g. 325 -> 399 through 32 measure-code fields) collapses to a single
        # labeled edge instead of stacking 32 identical arcs.
        rel_by_dir = {}
        for r in data["relationships"]:
            if r["target"] in form_names and r["source"] in form_names:
                rel_by_dir.setdefault((r["source"], r["target"]), []).append(r["via"])
        for (src, tgt), vias in rel_by_dir.items():
            label = vias[0] if len(vias) == 1 else f"{len(vias)} relationships"
            relationships.append({"source": fid(slug, src),
                                  "target": fid(slug, tgt), "label": label})
        for w in data["workflows"]:
            if not w["trigger"]:
                continue
            wid = f"WF::{slug}::{w['callsign']}"
            workflows.append({"id": wid, "callsign": w["callsign"], "name": w["name"],
                              "slug": slug, "enabled": w.get("enabled", True)})
            if w["trigger"]["form"] in form_names:
                wf_edges.append({"source": fid(slug, w["trigger"]["form"]), "target": wid,
                                 "label": "trigger"})
            for a in w["actions"]:
                if a["targetForm"] in form_names:
                    wf_edges.append({"source": wid, "target": fid(slug, a["targetForm"]),
                                     "label": a["type"].replace("BuiltIn.", "")})

    # Duplicate-form links chain each instance of a shared name across workspaces.
    dup_forms = []
    by_name = {}
    for f in forms:
        by_name.setdefault(f["name"], []).append(f)
    for name, instances in by_name.items():
        if len({i["slug"] for i in instances}) >= 2:
            dup_forms.append({"name": name, "ids": [i["id"] for i in sorted(instances, key=lambda x: x["slug"])]})

    viz = {
        "workspaces": [{"slug": w["slug"], "name": w["name"]} for w in agg["workspaces"]],
        "forms": forms,
        "workflows": workflows,
        "relationships": relationships,
        "wfEdges": wf_edges,
        "dupForms": dup_forms,
        "dupFlows": [{"signature": d["signature"], "members": d["members"]}
                     for d in agg["duplicateFlows"]],
        "stats": {
            "workspaces": len(agg["workspaces"]),
            "forms": len(forms),
            "workflows": len(workflows),
            "dupForms": len(dup_forms),
            "dupFlows": len(agg["duplicateFlows"]),
        },
    }
    data_json = json.dumps(viz, separators=(",", ":")).replace("</", "<\\/")
    html = TEMPLATE.replace("__DATA__", data_json)

    GLOBAL_DIR.mkdir(parents=True, exist_ok=True)
    out = GLOBAL_DIR / "global-explorer.html"
    out.write_text(html, encoding="utf-8")
    print(f"  Saved -> {out.relative_to(OUTPUT_DIR.parent)}  ({out.stat().st_size/1024:.1f} KB)")


def build():
    agg = aggregate()
    print(f"  Workspaces: {len(agg['workspaces'])}  Forms: {len(agg['allForms'])}  "
          f"Workflows: {len(agg['allWorkflows'])}  "
          f"Collisions: {len(agg['collisions'])}  DuplicateFlows: {len(agg['duplicateFlows'])}")
    _build_excel(agg)
    _build_html(agg)


if __name__ == "__main__":
    build()
